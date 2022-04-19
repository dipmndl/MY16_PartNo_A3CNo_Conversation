import numpy as np
import Formatting_Report_old as Fr
import pandas as pd
import numpy as np
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import load_workbook
from string import ascii_uppercase
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills

try:
    display_obj = Fr.FormatData.get_display_doors(3000, 1000)
    result_df = pd.read_excel(
        "D:\\2.0 HONDA\\Automation_Dev\\PartNo_A3CNo_Conversation\\Result\\CONF_3A0A_Rel_5.5_MP_AfterUpdate_13-04-2022.xlsx",
        sheet_name="Result")
    result_df.rename({"Unnamed: 2": "a"}, axis="columns", inplace=True)
    result_df.drop(["a"], axis=1, inplace=True)
    df1 = pd.DataFrame([[np.nan] * len(result_df.columns)], columns=result_df.columns)
    result_df = df1.append(result_df, ignore_index=True)
    dataframe = pd.DataFrame(result_df).set_index(["SrNo.", "ECN_New", "A3C_No"])

    with ExcelWriter("D:\\2.0 HONDA\\Automation_Dev\\PartNo_A3CNo_Conversation\\Result\\format_file_FP.xlsx") as writer:
        dataframe.to_excel(writer)

    workbook = load_workbook(
        filename="D:\\2.0 HONDA\\Automation_Dev\\PartNo_A3CNo_Conversation\\Result\\format_file_FP.xlsx")
    sheets = workbook.sheetnames
    ws = workbook[sheets[0]]
    ws.insert_cols(5)
    ws.insert_cols(8)
    ws.cell(row=1, column=4).value = 'Object_ID'
    ws.cell(row=1, column=5).value = 'New Difference'
    ws.cell(row=1, column=6).value = 'New_Data'
    ws.cell(row=1, column=7).value = 'Old_Data'
    ws.cell(row=1, column=8).value = 'Old Difference'


    for x in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        ws.merge_cells(start_row=1, start_column=x, end_row=2, end_column=x)

    ws["A1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["B1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["C1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["D1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["E1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["F1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["G1"].alignment = Alignment(horizontal='left', vertical='top')
    ws["H1"].alignment = Alignment(horizontal='left', vertical='top')

    ws['E1'].alignment = Alignment(wrap_text=True)
    ws['H1'].alignment = Alignment(wrap_text=True)

    # Column width
    for column in ascii_uppercase:
        if column == 'A':
            ws.column_dimensions[column].width = 5
            ws.column_dimensions[column].font = Font(bold=False)
        elif column == 'B':
            ws.column_dimensions[column].width = 22
        elif column == 'C':
            ws.column_dimensions[column].width = 20
        elif column == 'D':
            ws.column_dimensions[column].width = 17
        elif column == 'E':
            ws.column_dimensions[column].width = 9
        elif column == 'F':
            ws.column_dimensions[column].width = 32
        elif column == 'G':
            ws.column_dimensions[column].width = 32
        elif column == 'H':
            ws.column_dimensions[column].width = 9
        elif column == 'J':
            ws.column_dimensions[column].width = 26
        else:
            ws.column_dimensions[column].width = 12

    # Header Fill with Orange
    orangeFill = PatternFill(start_color='ffc000',
                             end_color='ffc000',
                             fill_type='solid')

    ws['A1'].fill = orangeFill
    ws['B1'].fill = orangeFill
    ws['C1'].fill = orangeFill
    ws['D1'].fill = orangeFill
    ws['E1'].fill = orangeFill
    ws['F1'].fill = orangeFill
    ws['G1'].fill = orangeFill
    ws['H1'].fill = orangeFill
    ws['I1'].fill = orangeFill
    ws['J1'].fill = orangeFill

    # define border formats

    thin_border = Border(left=Side(border_style='dashed', color='FF000000'),
                         right=Side(border_style='dashed', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
                         )

    thick_border = Border(left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          top=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='medium', color='FF000000')
                          )

    # Define fill format
    fill_cell = PatternFill(fill_type=fills.FILL_SOLID, start_color='92d050', end_color='92d050')
    fill_cell_grey = PatternFill(fill_type=fills.FILL_SOLID, start_color='808080', end_color='808080')
    fill_cell_generic = PatternFill(fill_type=fills.FILL_SOLID, start_color='ffe4e1', end_color='ffe4e1')
    fillfont = Font(size=23, underline='single', color='FFBB00', bold=True,
                    italic=True)  # We apply the following parameters to the text: size - 23, underline, color = FFBB00 (text color is specified in RGB), bold, oblique. If we do not need a bold font, we use the construction: bold = False. We act similarly if we do not need an oblique font: italic = False.

    # define size of the table
    row_num = ws.max_row - 1
    col_num = ws.max_column - 1
    # location of the Table
    row_loc = 1
    col_loc = 1

    # Number of Tables
    Table_num = 1
    dis = 0  # distance between the tables
    for _ in range(Table_num):
        for i in range(row_loc, row_loc + row_num):
            new_dr_val, old_dr_val = '', ''
            for j in range(col_loc, col_num + col_loc):
                ws.cell(row=i + 1, column=j + 1).border = thin_border
                if ws.cell(row=i + 1, column=j + 1).value is not None:
                    if '388' in ws.cell(row=i + 1, column=j + 1).value:
                        ws.cell(row=i + 1, column=j + 1).fill = fill_cell  # Green color in ECM_New column
                if j in [6]:
                    new_dr_val, old_dr_val, val = '', '', ''
                    if ws.cell(row=i + 1, column=j).value is not None:  # col = F
                        new_dr_val = ws.cell(row=i + 1, column=j).value  # new door value of col = F
                        if ws.cell(row=i + 1, column=j + 1).value is not None:  # col = G same row
                            old_dr_val = ws.cell(row=i + 1, column=j + 1).value  # old door value of col = G
                            if new_dr_val == old_dr_val:  # col F
                                ws.cell(row=i + 1, column=j + 3).fill = fill_cell_grey
                                ws.cell(row=i + 1, column=j + 4).fill = fill_cell_grey
                            else:
                                diff_val_new = Fr.find_difference(new_dr_val, old_dr_val, 'N')
                                diff_val_old = Fr.find_difference(new_dr_val, old_dr_val, 'O')
                                ws.cell(row=i + 1, column=j - 1).value = diff_val_new
                                ws.cell(row=i + 1, column=j + 2).value = diff_val_old
                                ws.cell(row=i + 1, column=j + 3).value = 'Ok'
                                ws.cell(row=i + 1, column=j + 4).value = 'This difference is Ok'
                                ws.cell(row=i + 1, column=j + 3).fill = fill_cell
                if i == row_loc + row_num - 1:
                    ws.cell(row=i + 1, column=j + 1).border = thick_border
        row_loc = row_loc + row_num + dis

    workbook.save(filename="D:\\2.0 HONDA\\Automation_Dev\\PartNo_A3CNo_Conversation\\Result\\CONF_3A0A_Rel_5.5_MP_Report_AfterUpdate.xlsx")


except Exception as e:
    print("Oops!", e.__class__, "occurred.")
    print(e)
